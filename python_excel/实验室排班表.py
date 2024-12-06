import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

# 创建一个Workbook和Worksheet
wb = Workbook()
ws = wb.active
ws.title = "卫生排班表"

# 设置标题和表头
title = "实验室卫生排班表"
ws['A1'] = title
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = f"月份：{datetime.now().month}"
ws['A3'] = f"年份：{datetime.now().year}"

# 设置表头
headers = ["日期", "星期", "值日生", "卫生检查员", "备注"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="00CCCC", end_color="00CCCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# 获取当前月份的第一天和最后一天
current_month = datetime.now().month
current_year = datetime.now().year
first_day = datetime(current_year, current_month, 1)
last_day = first_day + timedelta(days=32)  # 从1号开始加32天
last_day = last_day.replace(day=1) - timedelta(days=1)  # 减去1天得到最后一天

# 填充日期和星期
current_date = first_day
while current_date.day <= last_day.day:
    row_num = current_date.day - first_day.day + 5  # 计算行号
    ws.cell(row=row_num, column=1, value=current_date.day)
    ws.cell(row=row_num, column=2, value=current_date.strftime('%A'))  # 星期
    current_date += timedelta(days=1)

# 假设实验室有5个成员，轮流值日和检查
members = ["张三", "李四", "王五", "赵六", "孙七"]
for row_num in range(5, last_day.day - first_day.day + 6):  # 计算行数范围
    for col_num in range(3, 5):  # 值日生和卫生检查员
        member_index = (row_num - 5) % len(members)
        ws.cell(row=row_num, column=col_num, value=members[member_index])

# 保存Excel文件
wb.save("实验室卫生排班表.xlsx")